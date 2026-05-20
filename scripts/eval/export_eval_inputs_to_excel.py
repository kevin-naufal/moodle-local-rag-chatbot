from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def load_question_rows(path: Path) -> list[dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    questions = payload.get("questions") if isinstance(payload, dict) else payload
    if not isinstance(questions, list):
        raise ValueError("Question dataset must be a JSON array or an object with a 'questions' list.")

    source_document = str(payload.get("source_document") or "") if isinstance(payload, dict) else ""
    default_scope = str(payload.get("scope") or "") if isinstance(payload, dict) else ""
    batch = payload.get("batch") if isinstance(payload, dict) else ""

    rows: list[dict] = []
    for index, item in enumerate(questions, start=1):
        if not isinstance(item, dict):
            continue
        gold_points = item.get("gold_points") if isinstance(item.get("gold_points"), list) else []
        gold_source = item.get("gold_source") if isinstance(item.get("gold_source"), list) else []
        rows.append(
            {
                "row_no": index,
                "question_id": str(item.get("id") or item.get("question_id") or "").strip(),
                "question": str(item.get("question") or "").strip(),
                "scope": str(item.get("scope") or default_scope or "").strip(),
                "batch": batch,
                "source_document": source_document,
                "gold_points_count": len(gold_points),
                "gold_points": "\n".join(f"- {point}" for point in gold_points),
                "gold_source_count": len(gold_source),
                "gold_source": "\n".join(f"- {source}" for source in gold_source),
            }
        )
    return rows


def load_answer_run_rows(path: Path) -> list[dict]:
    rows: list[dict] = []
    with path.open(encoding="utf-8") as handle:
        for index, line in enumerate(handle, start=1):
            text = line.strip()
            if not text:
                continue
            item = json.loads(text)
            if not isinstance(item, dict):
                continue

            retrieved_context = item.get("retrieved_context") if isinstance(item.get("retrieved_context"), list) else []
            context_lines: list[str] = []
            for context_index, context_item in enumerate(retrieved_context, start=1):
                if not isinstance(context_item, dict):
                    continue
                source = str(context_item.get("source") or "").strip()
                page = context_item.get("page")
                text_value = str(context_item.get("text") or "").strip()
                prefix = f"[{context_index}] {source}"
                if page not in (None, ""):
                    prefix += f" p.{page}"
                context_lines.append(f"{prefix}\n{text_value}")

            corpus_sources = item.get("corpus_sources") if isinstance(item.get("corpus_sources"), list) else []
            rows.append(
                {
                    "row_no": index,
                    "question_id": str(item.get("question_id") or "").strip(),
                    "question": str(item.get("question") or "").strip(),
                    "mode": str(item.get("mode") or "").strip(),
                    "run_id": item.get("run_id"),
                    "model_name": str(item.get("model_name") or "").strip(),
                    "embedding_backend": str(item.get("embedding_backend") or "").strip(),
                    "status": str(item.get("status") or "").strip(),
                    "latency_total": item.get("latency_total"),
                    "latency_retrieval": item.get("latency_retrieval"),
                    "latency_generation": item.get("latency_generation"),
                    "timestamp": str(item.get("timestamp") or "").strip(),
                    "error_message": str(item.get("error_message") or "").strip(),
                    "model_answer": str(item.get("model_answer") or "").strip(),
                    "retrieved_context_count": len(retrieved_context),
                    "retrieved_context": "\n\n".join(context_lines),
                    "corpus_sources": "\n".join(f"- {source}" for source in corpus_sources),
                    "corpus_signature": str(item.get("corpus_signature") or "").strip(),
                    "corpus_data_dir": str(item.get("corpus_data_dir") or "").strip(),
                }
            )
    return rows


def style_worksheet(worksheet) -> None:
    worksheet.freeze_panes = "A2"
    header_font = Font(bold=True)
    top_alignment = Alignment(vertical="top", wrap_text=True)

    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = top_alignment

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = top_alignment

    for column_cells in worksheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        max_length = max(len(value) for value in values) if values else 10
        adjusted = min(max(max_length + 2, 10), 60)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted


def write_excel(rows: list[dict], output_path: Path, sheet_name: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    dataframe = pd.DataFrame(rows)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.book[sheet_name]
        style_worksheet(worksheet)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export evaluation question dataset and raw answer runs to Excel.")
    parser.add_argument("--questions", required=True, help="Path to question dataset JSON.")
    parser.add_argument("--answer-runs", required=True, help="Path to raw answer-runs JSONL.")
    parser.add_argument("--output-dir", default="data/excel_exports", help="Directory for generated Excel files.")
    args = parser.parse_args()

    project_root = Path(__file__).resolve().parents[2]
    questions_path = (project_root / args.questions).resolve() if not Path(args.questions).is_absolute() else Path(args.questions).resolve()
    answer_runs_path = (project_root / args.answer_runs).resolve() if not Path(args.answer_runs).is_absolute() else Path(args.answer_runs).resolve()
    output_dir = (project_root / args.output_dir).resolve() if not Path(args.output_dir).is_absolute() else Path(args.output_dir).resolve()

    question_rows = load_question_rows(questions_path)
    answer_run_rows = load_answer_run_rows(answer_runs_path)

    question_output = output_dir / f"{questions_path.stem}.xlsx"
    answer_run_output = output_dir / f"{answer_runs_path.stem}.xlsx"

    write_excel(question_rows, question_output, "questions")
    write_excel(answer_run_rows, answer_run_output, "answer_runs")

    print("Evaluation inputs exported to Excel.")
    print(f"- question_excel: {question_output}")
    print(f"- answer_runs_excel: {answer_run_output}")
    print(f"- question_rows: {len(question_rows)}")
    print(f"- answer_run_rows: {len(answer_run_rows)}")


if __name__ == "__main__":
    main()
